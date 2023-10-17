from imf import Series, Data
import pandas as pd
from openpyxl import Workbook

# https://datahelp.imf.org/knowledgebase/articles/1968408-how-to-use-the-api-python-and-r
# https://data.imf.org/

##### KOREA #####
# COUTRY_CODE = "KR"
# START_YEAR = 1995
# END_YEAR = 2022
# series = Series("Korea Data", START_YEAR, END_YEAR)
# series.add_data(
#     Data("Korea GDP", "Q", COUTRY_CODE, "NGDP_SA_XDC", START_YEAR, END_YEAR)
# )
# series.add_data(
#     Data("Korea USD rate", "Q", COUTRY_CODE, "ENDA_XDC_USD_RATE", START_YEAR, END_YEAR)
# )
# series.add_data(
#     Data("Korea policy rate", "Q", COUTRY_CODE, "FPOLM_PA", START_YEAR, END_YEAR)
# )
# series.draw_graph()


##### JAPAN #####
# COUTRY_CODE = "JP"
# START_YEAR = 1995
# END_YEAR = 2022
# japan_series = Series("Japan Data", START_YEAR, END_YEAR)
# japan_series.add_data(
#     Data("Japan GDP", "Q", COUTRY_CODE, "NGDP_SA_XDC", START_YEAR, END_YEAR)
# )
# japan_series.add_data(
#     Data("Japan USD rate", "Q", COUTRY_CODE, "ENDA_XDC_USD_RATE", START_YEAR, END_YEAR)
# )
# japan_series.add_data(
#     Data("Japan policy rate", "Q", COUTRY_CODE, "FPOLM_PA", START_YEAR, END_YEAR)
# )
# japan_series.draw_graph()


##### VIETNAM #####
COUTRY_CODE = "VN"
START_YEAR = 1995
END_YEAR = 2022
series = Series("Vietnam Data", START_YEAR, END_YEAR)
series.add_data(Data("Vietnam GDP", "A", COUTRY_CODE, "NGDP_XDC", START_YEAR, END_YEAR))
series.add_data(
    Data(
        "Vietnam USD rate", "M", COUTRY_CODE, "ENDA_XDC_USD_RATE", START_YEAR, END_YEAR
    )
)
series.add_data(
    Data("Vietnam policy rate", "Q", COUTRY_CODE, "FPOLM_PA", START_YEAR, END_YEAR)
)
series.draw_graph()


def convert_to_date(time):
    year = time // 12
    month = 1 + time % 12
    return f"{year}-{month:02d}"


for data in series.data_list:
    # 데이터프레임 생성
    df = pd.DataFrame(data.get_data())

    # time 열을 년도-월 형식으로 변환
    df["time"] = df["time"].apply(convert_to_date)

    # 새로운 엑셀 워크북 생성
    wb = Workbook()

    # 엑셀 시트 생성
    ws = wb.active

    # 데이터프레임의 열 이름을 엑셀 첫 번째 행에 쓰기
    for idx, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=idx, value=col)

    # 데이터프레임의 값들을 엑셀에 쓰기
    for r_idx, row in enumerate(df.itertuples(), start=2):
        for c_idx, value in enumerate(row[1:], start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 엑셀 파일 저장
    wb.save(data.title + "_economic_data.xlsx")
